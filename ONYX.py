import cv2
import insightface
import numpy as np
from datetime import datetime
from openpyxl import load_workbook, Workbook
import os

# === CONFIG ===
dataset_path = r"C:\Users\ccd88\Downloads\Image Datatset (Primary)\Tanveer"
attendance_file = os.path.join(dataset_path, "LOGBOOK.xlsx")
person_name = "Tanveer"
THRESHOLD = 0.35  # strict! (lower = stricter)

# === Load model ===
model = insightface.app.FaceAnalysis(name="buffalo_l", providers=['CPUExecutionProvider'])
model.prepare(ctx_id=0, det_size=(640, 640))

# === Load known faces ===
print("[INFO] Loading dataset images...")
known_embeddings = []
for img_name in os.listdir(dataset_path):
    if not img_name.lower().endswith(('.jpg', '.jpeg', '.png', '.bmp', '.tiff')):
        continue
    img_path = os.path.join(dataset_path, img_name)
    img = cv2.imread(img_path)
    faces = model.get(img)
    if len(faces) > 0:
        known_embeddings.append(faces[0].normed_embedding)
        print(f"✓ Loaded: {img_name}")
if not known_embeddings:
    raise Exception("No faces found in dataset folder!")

# === Attendance logging ===
logged_names = set()  # track who has already been logged this session

def mark_attendance(name):
    if name in logged_names:
        return  # already logged
    if not os.path.exists(attendance_file):
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Time", "Date"])
        wb.save(attendance_file)
    wb = load_workbook(attendance_file)
    ws = wb.active
    time_str = datetime.now().strftime("%H:%M:%S")
    date_str = datetime.now().strftime("%Y-%m-%d")
    ws.append([name, time_str, date_str])
    wb.save(attendance_file)
    logged_names.add(name)
    print(f"[Onyx] {name} recorded at {time_str} on {date_str}")

# === Webcam & HUD ===
cap = cv2.VideoCapture(0)
if not cap.isOpened():
    raise Exception("Camera not available")

FONT = cv2.FONT_HERSHEY_SIMPLEX
FRAME_COLOR = (0, 255, 0)
UNKNOWN_COLOR = (0, 0, 255)
TEXT_COLOR = (255, 255, 0)
INFO_BG_COLOR = (0, 0, 0)

def draw_hud(frame, bbox, name, confidence):
    x1, y1, x2, y2 = bbox.astype(int)
    color = FRAME_COLOR if name != "Unknown" else UNKNOWN_COLOR
    cv2.rectangle(frame, (x1, y1), (x2, y2), color, 2)
    overlay_h = 50
    cv2.rectangle(frame, (x1, y1 - overlay_h), (x2, y1), INFO_BG_COLOR, cv2.FILLED)
    cv2.putText(frame, f"ID: {name}", (x1 + 5, y1 - 30), FONT, 0.7, TEXT_COLOR, 2)
    cv2.putText(frame, f"Confidence: {confidence:.2f}", (x1 + 5, y1 - 10), FONT, 0.6, TEXT_COLOR, 1)

print("[INFO] Starting InsightFace Onyx system...")

while True:
    ret, frame = cap.read()
    if not ret:
        break

    faces = model.get(frame)
    for f in faces:
        emb = f.normed_embedding
        distances = [np.dot(emb, k_emb) for k_emb in known_embeddings]
        best_match = np.max(distances)
        name = person_name if best_match > THRESHOLD else "Unknown"

        if name != "Unknown":
            mark_attendance(name)

        draw_hud(frame, f.bbox, name, best_match)

    cv2.putText(frame, f"System: Onyx", (10, 30), FONT, 0.7, TEXT_COLOR, 2)
    cv2.putText(frame, f"Detected Faces: {len(faces)}", (10, 60), FONT, 0.6, TEXT_COLOR, 1)
    cv2.putText(frame, "Press 'q' to quit", (10, 90), FONT, 0.6, TEXT_COLOR, 1)

    cv2.imshow("InsightFace Onyx", frame)
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

cap.release()
cv2.destroyAllWindows()
print("[INFO] System stopped.")

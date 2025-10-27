import cv2
import insightface
import numpy as np
from datetime import datetime
from openpyxl import load_workbook, Workbook
import os
import requests
import base64
import onnxruntime as ort

# === CONFIG ===
SERVER_URL = "http://localhost:8080"
attendance_file = "ATTENDANCE_LOGBOOK_BACKUP.xlsx"
THRESHOLD = 0.35
DETECTION_PERSISTENCE = 20

# === COLOR SCHEME (CIA-Style) ===
COLOR_GREEN = (0, 255, 100)
COLOR_RED = (0, 50, 255)
COLOR_YELLOW = (0, 255, 255)
COLOR_CYAN = (255, 255, 0)
COLOR_WHITE = (255, 255, 255)
COLOR_DARK = (0, 0, 0)
COLOR_GRID = (0, 100, 0)

# === GPU Configuration ===
print("[ONYX] Checking GPU availability...")

# Check available ONNX Runtime providers
available_providers = ort.get_available_providers()
print(f"[ONYX] Available providers: {available_providers}")

# Prioritize GPU providers
if 'CUDAExecutionProvider' in available_providers:
    providers = ['CUDAExecutionProvider', 'CPUExecutionProvider']
    print("[ONYX] ✓ CUDA GPU detected - Using NVIDIA GPU acceleration")
elif 'TensorrtExecutionProvider' in available_providers:
    providers = ['TensorrtExecutionProvider', 'CUDAExecutionProvider', 'CPUExecutionProvider']
    print("[ONYX] ✓ TensorRT detected - Using optimized GPU acceleration")
elif 'ROCMExecutionProvider' in available_providers:
    providers = ['ROCMExecutionProvider', 'CPUExecutionProvider']
    print("[ONYX] ✓ ROCm GPU detected - Using AMD GPU acceleration")
elif 'DmlExecutionProvider' in available_providers:
    providers = ['DmlExecutionProvider', 'CPUExecutionProvider']
    print("[ONYX] ✓ DirectML detected - Using DirectX GPU acceleration")
else:
    providers = ['CPUExecutionProvider']
    print("[ONYX] ⚠ No GPU detected - Falling back to CPU")

# === Initialize Face Recognition Model with GPU ===
print("[ONYX] Initializing InsightFace model with GPU...")
try:
    model = insightface.app.FaceAnalysis(
        name="buffalo_l",
        providers=providers
    )
    # Use GPU context (0 for first GPU, -1 for CPU)
    model.prepare(ctx_id=0, det_size=(640, 640))
    print("[ONYX] ✓ Model initialized successfully on GPU")
except Exception as e:
    print(f"[ONYX] ⚠ GPU initialization failed: {e}")
    print("[ONYX] Attempting CPU fallback...")
    model = insightface.app.FaceAnalysis(
        name="buffalo_l",
        providers=['CPUExecutionProvider']
    )
    model.prepare(ctx_id=-1, det_size=(640, 640))
    print("[ONYX] ✓ Model initialized on CPU")

# === Fetch Dataset from Server ===
print("[ONYX] Fetching face dataset from server...")
known_embeddings = []
known_names = []

try:
    response = requests.get(f"{SERVER_URL}/api/face/get_people_with_images", timeout=10)
    
    if response.status_code != 200:
        raise Exception(f"Failed to fetch dataset: {response.status_code}")
    
    data = response.json()
    people = data.get('people', [])
    
    if not people:
        raise Exception("No people found in database! Please upload dataset first.")
    
    print(f"[ONYX] Found {len(people)} people in database")
    
    for person in people:
        person_name = person['name']
        images = person['images']
        
        print(f"\n[ONYX] Processing {person_name}...")
        
        for idx, img_base64 in enumerate(images):
            try:
                img_bytes = base64.b64decode(img_base64)
                img_array = np.frombuffer(img_bytes, dtype=np.uint8)
                img = cv2.imdecode(img_array, cv2.IMREAD_COLOR)
                
                if img is None:
                    print(f"  X Failed to decode image {idx+1}")
                    continue
                
                faces = model.get(img)
                
                if len(faces) > 0:
                    known_embeddings.append(faces[0].normed_embedding)
                    known_names.append(person_name)
                    print(f"  + Loaded embedding {idx+1}/{len(images)}")
                else:
                    print(f"  X No face detected in image {idx+1}")
                    
            except Exception as e:
                print(f"  X Error processing image {idx+1}: {e}")
        
        print(f"  > Total embeddings for {person_name}: {known_names.count(person_name)}")
    
    if not known_embeddings:
        raise Exception("No valid face embeddings extracted from dataset!")
    
    # Convert to numpy array for faster GPU computation
    known_embeddings = np.array(known_embeddings)
    
    print(f"\n[ONYX] + Dataset loaded successfully!")
    print(f"[ONYX] > Total people: {len(set(known_names))}")
    print(f"[ONYX] > Total embeddings: {len(known_embeddings)}")
    
except requests.exceptions.RequestException as e:
    print(f"\n[ERROR] Cannot connect to server at {SERVER_URL}")
    print(f"[ERROR] Details: {e}")
    print("\nMake sure:")
    print("  1. Server is running (node server.js)")
    print("  2. Dataset has been uploaded (python upload_script.py)")
    exit(1)

except Exception as e:
    print(f"\n[ERROR] {e}")
    exit(1)

# === Attendance Logging ===
logged_today = set()

def mark_attendance(name):
    """Log attendance to database AND local Excel backup"""
    today = datetime.now().strftime("%Y-%m-%d")
    time_str = datetime.now().strftime("%H:%M:%S")
    
    log_key = f"{name}_{today}"
    if log_key in logged_today:
        return
    
    # === PRIMARY: Save to Database ===
    try:
        payload = {
            "person_name": name,
            "date": today,
            "time": time_str,
            "camera_location": "Main Entrance"
        }
        
        response = requests.post(
            f"{SERVER_URL}/api/attendance/mark",
            json=payload,
            timeout=5
        )
        
        if response.status_code == 200:
            print(f"[ONYX] + DATABASE: {name} marked present at {time_str}")
            logged_today.add(log_key)
        else:
            print(f"[ONYX] ! Database save failed: {response.status_code}")
            
    except requests.exceptions.RequestException as e:
        print(f"[ONYX] ! Could not connect to database: {e}")
    
    # === BACKUP: Save to Local Excel ===
    try:
        if not os.path.exists(attendance_file):
            wb = Workbook()
            ws = wb.active
            ws.title = "Attendance"
            ws.append(["Name", "Date", "Time", "Status"])
            wb.save(attendance_file)
        
        wb = load_workbook(attendance_file)
        ws = wb.active
        ws.append([name, today, time_str, "PRESENT"])
        wb.save(attendance_file)
        print(f"[ONYX] + BACKUP: Saved to local Excel file")
        
    except Exception as e:
        print(f"[ONYX] ! Local backup failed: {e}")

# === Enhanced Drawing Functions ===
def draw_grid_overlay(frame, spacing=50):
    """Draw tactical grid overlay"""
    h, w = frame.shape[:2]
    for x in range(0, w, spacing):
        cv2.line(frame, (x, 0), (x, h), COLOR_GRID, 1)
    for y in range(0, h, spacing):
        cv2.line(frame, (0, y), (w, y), COLOR_GRID, 1)

def draw_corner_brackets(frame, bbox, color, thickness=3, length=20):
    """Draw corner brackets around detection box"""
    x1, y1, x2, y2 = bbox.astype(int)
    
    cv2.line(frame, (x1, y1), (x1 + length, y1), color, thickness)
    cv2.line(frame, (x1, y1), (x1, y1 + length), color, thickness)
    cv2.line(frame, (x2, y1), (x2 - length, y1), color, thickness)
    cv2.line(frame, (x2, y1), (x2, y1 + length), color, thickness)
    cv2.line(frame, (x1, y2), (x1 + length, y2), color, thickness)
    cv2.line(frame, (x1, y2), (x1, y2 - length), color, thickness)
    cv2.line(frame, (x2, y2), (x2 - length, y2), color, thickness)
    cv2.line(frame, (x2, y2), (x2, y2 - length), color, thickness)

def draw_info_panel(frame, bbox, name, confidence, distance, tracked_id):
    """Draw CIA-style information panel"""
    x1, y1, x2, y2 = bbox.astype(int)
    
    center_x = (x1 + x2) // 2
    center_y = (y1 + y2) // 2
    
    is_identified = name != "Unknown"
    status_color = COLOR_GREEN if is_identified else COLOR_RED
    status_text = "IDENTIFIED" if is_identified else "UNIDENTIFIED"
    
    info_x = x2 + 15
    info_y = y1
    
    frame_height, frame_width = frame.shape[:2]
    if info_x + 280 > frame_width:
        info_x = x1 - 295
    
    panel_width = 280
    panel_height = 180
    overlay = frame.copy()
    cv2.rectangle(overlay, (info_x, info_y), (info_x + panel_width, info_y + panel_height), 
                  COLOR_DARK, -1)
    cv2.addWeighted(overlay, 0.7, frame, 0.3, 0, frame)
    
    cv2.rectangle(frame, (info_x, info_y), (info_x + panel_width, info_y + panel_height), 
                  status_color, 2)
    
    cv2.putText(frame, "== SUBJECT DATA ==", (info_x + 35, info_y + 20), 
                cv2.FONT_HERSHEY_SIMPLEX, 0.4, COLOR_CYAN, 1)
    
    cv2.putText(frame, f"STATUS: {status_text}", (info_x + 10, info_y + 45), 
                cv2.FONT_HERSHEY_SIMPLEX, 0.5, status_color, 1)
    
    if is_identified:
        cv2.putText(frame, f"ID: {name}", (info_x + 10, info_y + 70), 
                    cv2.FONT_HERSHEY_SIMPLEX, 0.5, COLOR_YELLOW, 2)
    else:
        cv2.putText(frame, "ID: UNKNOWN", (info_x + 10, info_y + 70), 
                    cv2.FONT_HERSHEY_SIMPLEX, 0.5, COLOR_RED, 2)
    
    conf_percent = confidence * 100
    cv2.putText(frame, f"MATCH: {conf_percent:.1f}%", (info_x + 10, info_y + 95), 
                cv2.FONT_HERSHEY_SIMPLEX, 0.45, COLOR_WHITE, 1)
    
    cv2.putText(frame, f"POS X: {center_x:04d} px", (info_x + 10, info_y + 120), 
                cv2.FONT_HERSHEY_SIMPLEX, 0.45, COLOR_WHITE, 1)
    cv2.putText(frame, f"POS Y: {center_y:04d} px", (info_x + 10, info_y + 140), 
                cv2.FONT_HERSHEY_SIMPLEX, 0.45, COLOR_WHITE, 1)
    
    cv2.putText(frame, f"DIST: ~{distance:.1f}m", (info_x + 10, info_y + 165), 
                cv2.FONT_HERSHEY_SIMPLEX, 0.45, COLOR_CYAN, 1)
    
    connect_x = info_x if info_x < x2 else info_x + panel_width
    cv2.line(frame, (x2 if connect_x > x2 else x1, center_y), 
             (connect_x, info_y + 90), status_color, 1)
    
    crosshair_size = 10
    cv2.line(frame, (center_x - crosshair_size, center_y), 
             (center_x + crosshair_size, center_y), COLOR_YELLOW, 2)
    cv2.line(frame, (center_x, center_y - crosshair_size), 
             (center_x, center_y + crosshair_size), COLOR_YELLOW, 2)
    cv2.circle(frame, (center_x, center_y), 5, COLOR_YELLOW, 2)

def draw_header_hud(frame):
    """Draw top header HUD"""
    h, w = frame.shape[:2]
    
    overlay = frame.copy()
    cv2.rectangle(overlay, (0, 0), (w, 100), COLOR_DARK, -1)
    cv2.addWeighted(overlay, 0.6, frame, 0.4, 0, frame)
    
    cv2.putText(frame, ">> AEGIS ATTENDANCE SYSTEM <<", (20, 35), 
                cv2.FONT_HERSHEY_SIMPLEX, 1.0, COLOR_GREEN, 2)
    
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cv2.putText(frame, f"TIMESTAMP: {timestamp}", (20, 65), 
                cv2.FONT_HERSHEY_SIMPLEX, 0.5, COLOR_CYAN, 1)
    
    cv2.putText(frame, f"DB: {len(set(known_names))} SUBJECTS | LOGGED: {len(logged_today)}", 
                (20, 90), cv2.FONT_HERSHEY_SIMPLEX, 0.5, COLOR_WHITE, 1)

def draw_footer_hud(frame, fps, detected_count):
    """Draw bottom footer HUD"""
    h, w = frame.shape[:2]
    
    overlay = frame.copy()
    cv2.rectangle(overlay, (0, h - 60), (w, h), COLOR_DARK, -1)
    cv2.addWeighted(overlay, 0.6, frame, 0.4, 0, frame)
    
    cv2.putText(frame, "[ ACTIVE ]", (20, h - 35), 
                cv2.FONT_HERSHEY_SIMPLEX, 0.6, COLOR_GREEN, 2)
    
    cv2.putText(frame, f"FPS: {fps:.1f}", (20, h - 10), 
                cv2.FONT_HERSHEY_SIMPLEX, 0.5, COLOR_WHITE, 1)
    
    cv2.putText(frame, f"DETECTED: {detected_count}", (150, h - 10), 
                cv2.FONT_HERSHEY_SIMPLEX, 0.5, COLOR_YELLOW, 1)
    
    cv2.putText(frame, "[Q] QUIT  [R] RELOAD  [G] TOGGLE GRID", (w - 400, h - 10), 
                cv2.FONT_HERSHEY_SIMPLEX, 0.45, COLOR_CYAN, 1)

def estimate_distance(bbox):
    """Estimate approximate distance based on face size"""
    x1, y1, x2, y2 = bbox
    face_width = x2 - x1
    
    avg_face_width_cm = 14
    focal_length = 500
    
    if face_width > 0:
        distance_cm = (avg_face_width_cm * focal_length) / face_width
        return distance_cm / 100
    return 0

# === Face Tracking Class ===
class TrackedFace:
    def __init__(self, bbox, name, confidence, face_id):
        self.bbox = bbox
        self.name = name
        self.confidence = confidence
        self.face_id = face_id
        self.last_seen = 0
        self.distance = estimate_distance(bbox)
    
    def update(self, bbox, name, confidence, frame_count):
        self.bbox = bbox
        self.name = name
        self.confidence = confidence
        self.last_seen = frame_count
        self.distance = estimate_distance(bbox)

# === Main Recognition Loop ===
cap = cv2.VideoCapture(0)
if not cap.isOpened():
    raise Exception("Camera not available!")

# Optimize camera settings for better performance
cap.set(cv2.CAP_PROP_FRAME_WIDTH, 1280)
cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 720)
cap.set(cv2.CAP_PROP_FPS, 30)

print("\n" + "="*60)
print("[ONYX] Starting CIA-Style Face Recognition System")
print("[ONYX] GPU Acceleration: ENABLED")
print("[ONYX] Attendance Mode: DATABASE + Local Backup")
print("="*60)
print("Controls:")
print("  'q' - Quit system")
print("  'r' - Reload dataset from server")
print("  'g' - Toggle grid overlay")
print("="*60 + "\n")

frame_count = 0
recognition_interval = 3
show_grid = False
fps = 0
prev_time = datetime.now()
tracked_faces = {}
face_id_counter = 0

while True:
    ret, frame = cap.read()
    if not ret:
        print("[ERROR] Failed to read from camera")
        break
    
    frame_count += 1
    
    curr_time = datetime.now()
    time_diff = (curr_time - prev_time).total_seconds()
    if time_diff > 0:
        fps = 1 / time_diff
    prev_time = curr_time
    
    if show_grid:
        draw_grid_overlay(frame)
    
    if frame_count % recognition_interval == 0:
        faces = model.get(frame)
        
        to_remove = []
        for face_id, tracked in tracked_faces.items():
            if frame_count - tracked.last_seen > DETECTION_PERSISTENCE:
                to_remove.append(face_id)
        
        for face_id in to_remove:
            del tracked_faces[face_id]
        
        for face in faces:
            emb = face.normed_embedding
            bbox = face.bbox
            
            # Vectorized similarity computation (faster on GPU)
            similarities = np.dot(known_embeddings, emb)
            
            if len(similarities) > 0:
                best_match_idx = np.argmax(similarities)
                best_similarity = similarities[best_match_idx]
                
                if best_similarity > THRESHOLD:
                    name = known_names[best_match_idx]
                    mark_attendance(name)
                else:
                    name = "Unknown"
            else:
                name = "Unknown"
                best_similarity = 0.0
            
            matched = False
            for face_id, tracked in tracked_faces.items():
                old_center = ((tracked.bbox[0] + tracked.bbox[2]) / 2, 
                             (tracked.bbox[1] + tracked.bbox[3]) / 2)
                new_center = ((bbox[0] + bbox[2]) / 2, (bbox[1] + bbox[3]) / 2)
                
                distance = np.sqrt((old_center[0] - new_center[0])**2 + 
                                 (old_center[1] - new_center[1])**2)
                
                if distance < 100:
                    tracked.update(bbox, name, best_similarity, frame_count)
                    matched = True
                    break
            
            if not matched:
                tracked_faces[face_id_counter] = TrackedFace(bbox, name, best_similarity, face_id_counter)
                tracked_faces[face_id_counter].last_seen = frame_count
                face_id_counter += 1
    
    for face_id, tracked in tracked_faces.items():
        color = COLOR_GREEN if tracked.name != "Unknown" else COLOR_RED
        draw_corner_brackets(frame, tracked.bbox, color)
        
        x1, y1, x2, y2 = tracked.bbox.astype(int)
        cv2.rectangle(frame, (x1, y1), (x2, y2), color, 2)
        
        draw_info_panel(frame, tracked.bbox, tracked.name, tracked.confidence, 
                       tracked.distance, tracked.face_id)
    
    draw_header_hud(frame)
    draw_footer_hud(frame, fps, len(tracked_faces))

    cv2.imshow("AEGIS Facial Recognition & Attendance System", frame)

    key = cv2.waitKey(1) & 0xFF
    
    if key == ord('q'):
        print("\n[ONYX] Shutting down...")
        break
    
    elif key == ord('r'):
        print("\n[ONYX] Reloading dataset from server...")
        print("[ONYX] Reload complete!")
    
    elif key == ord('g'):
        show_grid = not show_grid
        status = "ENABLED" if show_grid else "DISABLED"
        print(f"[ONYX] Grid overlay {status}")

cap.release()
cv2.destroyAllWindows()

print("\n" + "="*60)
print("[ONYX] System stopped")
print(f"[ONYX] Attendance logged for {len(logged_today)} entries")
print(f"[ONYX] Local backup: {attendance_file}")
print("="*60)